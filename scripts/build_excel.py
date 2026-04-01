"""
build_excel.py — FSA Store Service Intelligence Excel workbook
--------------------------------------------------------------
Produces a date-stamped workbook from the enriched unified CSV.

Tab order:
  1. L1 Category Mix Binary        (Yes/No, non-oil)
  2. L2 Category Mix Binary        (Yes/No, non-oil)
  3. L1 Category Mix Trans Count   (raw counts, non-oil)
  4. L2 Category Mix Trans Count   (raw counts, non-oil)
  5. L1 Oil Incl. Binary           (Yes/No, oil-inclusive)
  6. Oil vs Non-Oil Discrepancy    (flagged where oil=Yes but non-oil=No)
  7. Brand Summary
  8. Model Summary
  9. Region Summary
 10. DMA Summary
 11. Hierarchy Reference
 12. Raw Data
 13. Store Directory
"""

import csv, sys, argparse
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed.  Run: pip install openpyxl")
    sys.exit(1)

# ── CATEGORY DEFINITIONS ───────────────────────────────────────────────────────
CAT_COLORS = {
    'Air Conditioning':        '0EA5E9',
    'Air Filters':             '16A34A',
    'Battery':                 'D97706',
    'Brakes':                  'DC2626',
    'Differentials':           '7C3AED',
    'Emissions & Inspections*':'4F46E5',   # asterisk = state program flag
    'Engine Maintenance':        '2563EB',
    'Fluids & Cooling':        '0891B2',
    'Fuel System':             'B45309',
    'Lighting':                'CA8A04',
    'Oil Change':              '374151',
    'Shop & Misc':             '6B7280',
    'Suspension & Steering':   '059669',
    'Tires':                   '65A30D',
    'Transmission':            '9333EA',
    'Wiper Blades':            '475569',
    'Additives':               'D97706',
}

AUTHORITATIVE_L2 = {
    'Air Conditioning':        ['A/C Service'],
    'Additives':               ['Engine Cleaning Service', 'Engine Treatment', 'Fuel System Cleaner',
                                'High Mileage Treatment', 'Oil System Cleaner', 'Stop Leak'],
    'Air Filters':             ['Cabin Air Filter', 'Engine Air Filter'],
    'Battery':                 ['Battery Fees', 'Battery Replacement', 'Battery Service', 'Starting & Charging'],
    'Brakes':                  ['Brake Labor', 'Brake Pads — Front', 'Brake Pads — Rear',
                                'Brake Pads (General)', 'Rotors — Front', 'Rotors — Rear', 'Rotors (General)'],
    'Differentials':           ['Front Differential', 'Gear Oil', 'Rear Differential'],
    'Emissions & Inspections*': ['Emissions Test', 'Vehicle Check / Inspection'],
    'Engine Maintenance':       ['Belts', 'Ball Joints', 'Catalytic Converter', 'Clutch / Cylinder',
                                'Engine Labor', 'Hose Replacement', 'Ignition / Coils',
                                'O2 / Sensors', 'Spark Plugs', 'Thermostat / Water Pump',
                                'Valve Cover Gasket', 'Wheel Bearings'],
    'Fluids & Cooling':        ['Brake Fluid', 'Coolant / Antifreeze', 'Cooling System Labor',
                                'Power Steering Fluid', 'Radiator Replacement', 'Window Wash'],
    'Fuel System':             ['Fuel Filter', 'Fuel System Cleaning (BG44K / Injector)'],
    'Lighting':                ['Headlight Bulbs', 'Headlight Restoration', 'Interior / Signal Bulbs'],
    'Oil Change':              ['Oil Change Service'],
    'Shop & Misc':             ['Car Wash', 'Lug Nut Service', 'Shop Labor', 'Shop Supplies'],
    'Suspension & Steering':   ['Suspension / Steering Labor', 'Tie Rods', 'Wheel Alignment'],
    'Tires':                   ['Tire Disposal', 'Tire Mount & Balance', 'Tire Repair',
                                'Tire Replacement', 'Tire Rotation', 'TPMS'],
    'Transmission':            ['Drivetrain Labor', 'Transfer Case', 'Transmission Fluid Exchange'],
    'Wiper Blades':            ['Front Wiper Blades', 'Rear Wipers'],
}

# Non-oil categories only (exclude Oil Change from binary/count views)
NON_OIL_CATS = [c for c in AUTHORITATIVE_L2 if c != 'Oil Change']
L1_CATS      = list(AUTHORITATIVE_L2.keys())   # all including Oil Change
ALL_MODELS   = ['QL', 'QL+', 'QL++', 'FS']
MODEL_COLORS = {'QL': '1D4ED8', 'QL+': '7C3AED', 'QL++': 'D97706', 'FS': 'DC2626'}
BRANDS       = sorted(['American Lubefast', 'Economy Oil Change', 'Grease Monkey',
                        'Grease Monkey (ADI)', 'Herbert Automotive', 'Kwik Kar',
                        'SpeeDee', "Uncle Ed's"])

FSA_RED = 'C8102E'; DARK = '1F2937'; DARK2 = '374151'
WHITE   = 'FFFFFF'; LGREY = 'F9FAFB'; MGREY = 'F3F4F6'
DGREY   = '6B7280'; BLACK = '111827'
YES_FILL = 'D1FAE5'; YES_FONT = '065F46'     # green Yes
NO_FILL  = 'FEF2F2'; NO_FONT  = '991B1B'     # red   No
FLAG_FILL= 'FEF3C7'; FLAG_FONT= '92400E'     # amber flag (discrepancy)

# ── STYLE HELPERS ──────────────────────────────────────────────────────────────
def fill(h):    return PatternFill('solid', start_color=h, end_color=h)
def font(bold=False, color='000000', size=11, italic=False):
    return Font(name='Arial', bold=bold, color=color, size=size, italic=italic)
def bdr():
    s = Side(style='thin', color='D1D5DB')
    return Border(top=s, bottom=s, left=s, right=s)
def align(h='left', v='center', wrap=False, rot=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, textRotation=rot)

def set_title(ws, text, n_cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = font(bold=True, color=WHITE, size=12)
    c.fill = fill(DARK); c.alignment = align('left', 'center')
    ws.row_dimensions[1].height = 26

def set_id_hdr(ws, row, col, text, width):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(bold=True, color=WHITE, size=11)
    c.fill = fill(DARK2); c.alignment = align('left', 'center'); c.border = bdr()
    ws.column_dimensions[get_column_letter(col)].width = width

def set_cat_hdr(ws, row, col, cat, rotation=60, width=11):
    color = CAT_COLORS.get(cat, '6B7280')
    c = ws.cell(row=row, column=col, value=cat)
    c.font = font(bold=True, color=WHITE, size=10)
    c.fill = fill(color); c.alignment = align('center', 'bottom', rot=rotation); c.border = bdr()
    ws.column_dimensions[get_column_letter(col)].width = width

def set_count_cell(ws, row, col, count, color_hex):
    c = ws.cell(row=row, column=col, value=count)
    c.font = font(bold=True, color=color_hex, size=10)
    c.fill = fill(MGREY); c.alignment = align('center', 'center'); c.border = bdr()

def set_count_label(ws, row, col_start, col_end, text='Stores with transactions →'):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.font = font(bold=True, color=DGREY, size=9, italic=True)
    c.fill = fill(MGREY); c.alignment = align('right', 'center'); c.border = bdr()
    ws.row_dimensions[row].height = 16

def hm_fill(pct):
    if not pct: return fill(WHITE)
    if pct >= 0.60: return fill('D1FAE5')
    if pct >= 0.20: return fill('FEF3C7')
    return fill('FEE2E2')
def hm_font(pct):
    if not pct: return font(color=DGREY, size=10)
    if pct >= 0.60: return font(bold=True, color='065F46', size=10)
    if pct >= 0.20: return font(color='92400E', size=10)
    return font(color='991B1B', size=10)
def hm_val(pct): return f'{pct:.0%}' if pct else '—'

# ── ID COLUMNS ─────────────────────────────────────────────────────────────────
# 8 left-side identifier columns used on all matrix tabs
ID_COLS = [
    ('Store #',      9,  'STORE_NUMBER'),
    ('Brand',        20, 'BRAND'),
    ('City, State',  20, None),          # computed: CITY + STATE
    ('DMA',          24, 'DMA_NAME'),
    ('Region',       8,  'REGION_NUM'),
    ('District',     9,  'DISTRICT_NUM'),
    ('District Mgr', 20, 'DISTRICT_MGR'),
    ('Region VP',    20, 'REGION_VP'),
]
N_ID = len(ID_COLS)   # 8

def write_id_hdrs(ws, row):
    for col, (hdr, w, _) in enumerate(ID_COLS, 1):
        set_id_hdr(ws, row, col, hdr, w)

def write_id_vals(ws, row, s, bg):
    for col, (_, _, key) in enumerate(ID_COLS, 1):
        if key is None:
            val = f"{s['CITY']}, {s['STATE']}"
        else:
            val = s.get(key, '')
        bold = (col == 2)   # bold brand
        c = ws.cell(row=row, column=col, value=val)
        c.font = font(size=10, bold=bold, color=BLACK)
        c.fill = fill(bg); c.alignment = align('left', 'center'); c.border = bdr()


# ── MAIN ───────────────────────────────────────────────────────────────────────
def build(input_path, output_path, region_district_path=None):
    print(f"Reading: {input_path}")
    with open(input_path, newline='', encoding='utf-8-sig') as f:
        rows = list(csv.DictReader(f))
    print(f"  {len(rows):,} rows")

    today         = datetime.now()
    date_window   = f"Dec 17, 2025 – {today.strftime('%b %d, %Y')}"

    # ── Region/district lookup ─────────────────────────────────────────────────
    region_dist = {}
    if region_district_path:
        try:
            with open(region_district_path) as f:
                for row in csv.DictReader(f):
                    region_dist[str(row['STORE_NUMBER']).strip()] = {
                        'REGION_NUM':   row.get('REGION_NUM','').strip(),
                        'REGION_VP':    row.get('REGION_VP','').strip(),
                        'DISTRICT_NUM': row.get('DISTRICT_NUM','').strip(),
                        'DISTRICT_MGR': row.get('DISTRICT_MGR','').strip(),
                    }
            print(f"  Loaded {len(region_dist)} region/district records")
        except FileNotFoundError:
            print("  WARNING: region_district.csv not found")

    # ── Build store map ────────────────────────────────────────────────────────
    store_map = {}
    for r in rows:
        sn = r['STORE_NUMBER']
        if sn not in store_map:
            rd = region_dist.get(sn, {})
            if not rd.get('REGION_NUM') and r.get('REGION_NUM'):
                rd = {'REGION_NUM':   r.get('REGION_NUM',''),
                      'REGION_VP':    r.get('REGION_VP',''),
                      'DISTRICT_NUM': r.get('DISTRICT_NUM',''),
                      'DISTRICT_MGR': r.get('DISTRICT_MGR','')}
            store_map[sn] = dict(
                STORE_NUMBER=sn, STORE_NAME=r['STORE_NAME'],
                BRAND=r.get('BRAND',''), CITY=r['CITY'], STATE=r['STATE'],
                ADDRESS1=r.get('ADDRESS1',''), ZIP_CODE=r.get('ZIP_CODE',''),
                DMA_NAME=r.get('DMA_NAME',''),
                STORE_MODEL=r.get('STORE_MODEL',''),
                PROPOSED_MODEL=r.get('PROPOSED_MODEL',''),
                REGION_NUM=rd.get('REGION_NUM',''),
                REGION_VP=rd.get('REGION_VP',''),
                DISTRICT_NUM=rd.get('DISTRICT_NUM',''),
                DISTRICT_MGR=rd.get('DISTRICT_MGR',''),
                # non-oil accumulators
                l1_tx=defaultdict(int), l2_tx=defaultdict(int),
                # oil-inclusive accumulators
                l1_tx_all=defaultdict(int), l2_tx_all=defaultdict(int),
            )
        if r.get('OFFERS_SERVICE') == '1':
            is_oil = str(r.get('IS_OIL_TRANSACTION','0')).strip() == '1'
            cat = r['L1_CATEGORY']; sub = r.get('L2_SUBCATEGORY','')
            try: tx = int(r['TRANSACTION_COUNT'])
            except: tx = 0
            store_map[sn]['l1_tx_all'][cat] += tx
            store_map[sn]['l2_tx_all'][(cat, sub)] += tx
            if not is_oil:
                store_map[sn]['l1_tx'][cat] += tx
                store_map[sn]['l2_tx'][(cat, sub)] += tx

    stores = sorted(store_map.values(),
                    key=lambda x: (x['REGION_NUM'] or '9', x['DISTRICT_NUM'] or '999',
                                   x['BRAND'], x['STORE_NAME']))
    n_stores = len(stores)
    print(f"  {n_stores} unique stores")

    # ── Aggregate stats (non-oil, for summary tabs) ───────────────────────────
    l1_stats = defaultdict(lambda: {'stores': set(), 'tx': 0})
    l2_stats = defaultdict(lambda: {'stores': set(), 'tx': 0})
    brand_stores = defaultdict(set); brand_l1 = defaultdict(lambda: defaultdict(set))
    model_stores = defaultdict(set); model_l1 = defaultdict(lambda: defaultdict(set))
    region_stores= defaultdict(set); region_l1= defaultdict(lambda: defaultdict(set))
    dma_stores   = defaultdict(set); dma_l1   = defaultdict(lambda: defaultdict(set))

    for r in rows:
        if r.get('OFFERS_SERVICE') == '1' and str(r.get('IS_OIL_TRANSACTION','0')).strip() != '1':
            l1_stats[r['L1_CATEGORY']]['stores'].add(r['STORE_NUMBER'])
            l1_stats[r['L1_CATEGORY']]['tx'] += int(r.get('TRANSACTION_COUNT',0) or 0)
            if r.get('L2_SUBCATEGORY'):
                l2_stats[(r['L1_CATEGORY'],r['L2_SUBCATEGORY'])]['stores'].add(r['STORE_NUMBER'])
                l2_stats[(r['L1_CATEGORY'],r['L2_SUBCATEGORY'])]['tx'] += int(r.get('TRANSACTION_COUNT',0) or 0)

    for s in stores:
        brand_stores[s['BRAND']].add(s['STORE_NUMBER'])
        model_stores[s['STORE_MODEL']].add(s['STORE_NUMBER'])
        dma_stores[s['DMA_NAME']].add(s['STORE_NUMBER'])
        rn = s['REGION_NUM'] or 'Unknown'
        region_stores[rn].add(s['STORE_NUMBER'])
        for cat in s['l1_tx']:
            brand_l1[s['BRAND']][cat].add(s['STORE_NUMBER'])
            model_l1[s['STORE_MODEL']][cat].add(s['STORE_NUMBER'])
            dma_l1[s['DMA_NAME']][cat].add(s['STORE_NUMBER'])
            region_l1[rn][cat].add(s['STORE_NUMBER'])

    DMAS    = sorted(dma_stores.keys())
    REGIONS = sorted(region_stores.keys(), key=lambda x: int(x) if x.isdigit() else 999)
    REGION_LABELS = {s['REGION_NUM']: s['REGION_VP'] for s in stores if s['REGION_NUM']}
    all_l2  = [(l1, l2) for l1 in NON_OIL_CATS for l2 in AUTHORITATIVE_L2[l1]]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ══════════════════════════════════════════════════════════════════════════
    # Helper: build a binary matrix sheet (Yes/No)
    # ══════════════════════════════════════════════════════════════════════════
    def build_binary_l1(sheet_name, title, cats, get_tx, store_count_fn=None):
        ws = wb.create_sheet(sheet_name)
        n_c = len(cats)
        ws.freeze_panes = get_column_letter(N_ID+1) + '4'
        set_title(ws, title, N_ID + n_c)
        write_id_hdrs(ws, 2)
        for col, cat in enumerate(cats, N_ID+1):
            set_cat_hdr(ws, 2, col, cat, rotation=60)
        ws.row_dimensions[2].height = 80

        # Row 3: store counts with transactions
        set_count_label(ws, 3, 1, N_ID, 'Stores with transactions →')
        for col, cat in enumerate(cats, N_ID+1):
            cnt = store_count_fn(cat) if store_count_fn else len(l1_stats[cat]['stores'])
            set_count_cell(ws, 3, col, cnt, CAT_COLORS.get(cat,'6B7280'))

        for i, s in enumerate(stores, 4):
            bg = LGREY if i % 2 == 0 else WHITE
            write_id_vals(ws, i, s, bg)
            for col, cat in enumerate(cats, N_ID+1):
                tx = get_tx(s, cat)
                yes = tx > 0
                c = ws.cell(row=i, column=col, value='Yes' if yes else 'No')
                c.fill = fill(YES_FILL if yes else NO_FILL)
                c.font = font(bold=yes, color=YES_FONT if yes else NO_FONT, size=10)
                c.alignment = align('center','center'); c.border = bdr()
            ws.row_dimensions[i].height = 15

        ws.auto_filter.ref = f'A2:{get_column_letter(N_ID+n_c)}{3+n_stores}'
        return ws

    def build_count_l1(sheet_name, title, cats, get_tx):
        ws = wb.create_sheet(sheet_name)
        n_c = len(cats)
        ws.freeze_panes = get_column_letter(N_ID+1) + '4'
        set_title(ws, title, N_ID + n_c)
        write_id_hdrs(ws, 2)
        for col, cat in enumerate(cats, N_ID+1):
            set_cat_hdr(ws, 2, col, cat, rotation=60)
        ws.row_dimensions[2].height = 80
        set_count_label(ws, 3, 1, N_ID, 'Stores with transactions →')
        for col, cat in enumerate(cats, N_ID+1):
            cnt = len(l1_stats[cat]['stores'])
            set_count_cell(ws, 3, col, cnt, CAT_COLORS.get(cat,'6B7280'))
        for i, s in enumerate(stores, 4):
            bg = LGREY if i % 2 == 0 else WHITE
            write_id_vals(ws, i, s, bg)
            for col, cat in enumerate(cats, N_ID+1):
                tx = get_tx(s, cat)
                c = ws.cell(row=i, column=col, value=tx if tx else None)
                c.fill = fill('ECFDF5') if tx else fill(bg)
                c.font = font(bold=bool(tx), color='065F46' if tx else DGREY, size=10)
                c.alignment = align('center','center'); c.border = bdr()
            ws.row_dimensions[i].height = 15
        ws.auto_filter.ref = f'A2:{get_column_letter(N_ID+n_c)}{3+n_stores}'
        return ws

    def build_binary_l2(sheet_name, title, pairs, get_tx_l2, store_count_fn_l2=None):
        ws = wb.create_sheet(sheet_name)
        n_c = len(pairs)
        ws.freeze_panes = get_column_letter(N_ID+1) + '5'
        set_title(ws, title, N_ID + n_c)

        # Row 2: L1 group headers (merged)
        for col in range(1, N_ID+1):
            ws.cell(row=2, column=col).fill = fill(DARK2)
            ws.cell(row=2, column=col).border = bdr()
        col = N_ID+1
        for l1 in NON_OIL_CATS:
            l2s = AUTHORITATIVE_L2[l1]; color = CAT_COLORS.get(l1,'6B7280')
            if len(l2s) > 1:
                ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+len(l2s)-1)
            c = ws.cell(row=2, column=col, value=l1)
            c.font = font(bold=True, color=WHITE, size=10)
            c.fill = fill(color); c.alignment = align('center','center'); c.border = bdr()
            col += len(l2s)
        ws.row_dimensions[2].height = 18

        # Row 3: L2 subcategory headers
        write_id_hdrs(ws, 3)
        for col, (l1, l2) in enumerate(pairs, N_ID+1):
            color = CAT_COLORS.get(l1,'6B7280')
            c = ws.cell(row=3, column=col, value=l2)
            c.font = font(bold=True, color=WHITE, size=10)
            c.fill = fill(color); c.alignment = align('center','bottom',rot=60); c.border = bdr()
            ws.column_dimensions[get_column_letter(col)].width = 9
        ws.row_dimensions[3].height = 80

        # Row 4: store counts
        set_count_label(ws, 4, 1, N_ID, 'Stores with transactions →')
        for col, (l1, l2) in enumerate(pairs, N_ID+1):
            cnt = store_count_fn_l2(l1,l2) if store_count_fn_l2 else len(l2_stats[(l1,l2)]['stores'])
            set_count_cell(ws, 4, col, cnt, CAT_COLORS.get(l1,'6B7280'))

        for i, s in enumerate(stores, 5):
            bg = LGREY if i % 2 == 0 else WHITE
            write_id_vals(ws, i, s, bg)
            for col, (l1, l2) in enumerate(pairs, N_ID+1):
                tx = get_tx_l2(s, l1, l2)
                yes = tx > 0
                c = ws.cell(row=i, column=col, value='Yes' if yes else 'No')
                c.fill = fill(YES_FILL if yes else NO_FILL)
                c.font = font(bold=yes, color=YES_FONT if yes else NO_FONT, size=10)
                c.alignment = align('center','center'); c.border = bdr()
            ws.row_dimensions[i].height = 15

        ws.auto_filter.ref = f'A3:{get_column_letter(N_ID+n_c)}{4+n_stores}'
        return ws

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1: L1 Category Mix Binary (non-oil)
    # ══════════════════════════════════════════════════════════════════════════
    build_binary_l1(
        'L1 Category Mix Binary',
        f'FullSpeed Automotive — Service Category Mix (Yes/No)  |  Yes = ≥1 transaction recorded  |  {date_window}  |  *Excl. oil change  |  Emissions* = state inspection programs only (CO/WA/GA/TX/NC/UT)',
        NON_OIL_CATS,
        lambda s, cat: s['l1_tx'].get(cat, 0)
    )

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2: L2 Category Mix Binary (non-oil)
    # ══════════════════════════════════════════════════════════════════════════
    build_binary_l2(
        'L2 Category Mix Binary',
        f'FullSpeed Automotive — Service Subcategory Mix (Yes/No)  |  {date_window}  *Excluding Oil Change Transactions*',
        all_l2,
        lambda s, l1, l2: s['l2_tx'].get((l1,l2), 0)
    )

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3: L1 Category Mix Trans Count (non-oil)
    # ══════════════════════════════════════════════════════════════════════════
    build_count_l1(
        'L1 Category Mix Trans Count',
        f'FullSpeed Automotive — Service Transaction Counts by Store  |  Values = transaction count  |  Blank = 0 transactions  |  {date_window}',
        NON_OIL_CATS,
        lambda s, cat: s['l1_tx'].get(cat, 0)
    )

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4: L2 Category Mix Trans Count (non-oil)
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet('L2 Category Mix Trans Count')
    n_c4 = len(all_l2)
    ws4.freeze_panes = get_column_letter(N_ID+1) + '5'
    set_title(ws4, f'FullSpeed Automotive — Subcategory Transaction Counts  |  {date_window}', N_ID+n_c4)
    for col in range(1, N_ID+1):
        ws4.cell(row=2, column=col).fill = fill(DARK2); ws4.cell(row=2, column=col).border = bdr()
    col = N_ID+1
    for l1 in NON_OIL_CATS:
        l2s = AUTHORITATIVE_L2[l1]; color = CAT_COLORS.get(l1,'6B7280')
        if len(l2s) > 1:
            ws4.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+len(l2s)-1)
        c = ws4.cell(row=2, column=col, value=l1)
        c.font = font(bold=True, color=WHITE, size=10)
        c.fill = fill(color); c.alignment = align('center','center'); c.border = bdr()
        col += len(l2s)
    ws4.row_dimensions[2].height = 18
    write_id_hdrs(ws4, 3)
    for col, (l1, l2) in enumerate(all_l2, N_ID+1):
        color = CAT_COLORS.get(l1,'6B7280')
        c = ws4.cell(row=3, column=col, value=l2)
        c.font = font(bold=True, color=WHITE, size=10)
        c.fill = fill(color); c.alignment = align('center','bottom',rot=60); c.border = bdr()
        ws4.column_dimensions[get_column_letter(col)].width = 9
    ws4.row_dimensions[3].height = 80
    set_count_label(ws4, 4, 1, N_ID, 'Stores with transactions →')
    for col, (l1, l2) in enumerate(all_l2, N_ID+1):
        set_count_cell(ws4, 4, col, len(l2_stats[(l1,l2)]['stores']), CAT_COLORS.get(l1,'6B7280'))
    for i, s in enumerate(stores, 5):
        bg = LGREY if i % 2 == 0 else WHITE
        write_id_vals(ws4, i, s, bg)
        for col, (l1, l2) in enumerate(all_l2, N_ID+1):
            tx = s['l2_tx'].get((l1,l2), 0)
            c = ws4.cell(row=i, column=col, value=tx if tx else None)
            c.fill = fill('ECFDF5') if tx else fill(bg)
            c.font = font(bold=bool(tx), color='065F46' if tx else DGREY, size=10)
            c.alignment = align('center','center'); c.border = bdr()
        ws4.row_dimensions[i].height = 15
    ws4.auto_filter.ref = f'A3:{get_column_letter(N_ID+n_c4)}{4+n_stores}'

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 5: L1 Oil-Inclusive Binary
    # ══════════════════════════════════════════════════════════════════════════
    # Count stores with oil-inclusive transactions per category
    oil_l1_stores = defaultdict(set)
    for r in rows:
        if r.get('OFFERS_SERVICE') == '1':
            oil_l1_stores[r['L1_CATEGORY']].add(r['STORE_NUMBER'])

    build_binary_l1(
        'L1 Oil Incl. Binary',
        f'FullSpeed Automotive — Service Mix (Yes/No, Oil-Inclusive)  |  Includes services recorded during oil change visits  |  {date_window}',
        L1_CATS,
        lambda s, cat: s['l1_tx_all'].get(cat, 0),
        store_count_fn=lambda cat: len(oil_l1_stores[cat])
    )

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 6: Oil vs Non-Oil Discrepancy
    # Shows FLAG where oil-inclusive = Yes but non-oil = No
    # (service only appears when oil change transaction is present)
    # ══════════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet('Oil vs Non-Oil Discrepancy')
    DISC_CATS = NON_OIL_CATS   # Oil Change itself excluded — discrepancy only meaningful for other cats
    n_c6 = len(DISC_CATS)
    ws6.freeze_panes = get_column_letter(N_ID+1) + '4'
    set_title(ws6,
        f'FullSpeed Automotive — Oil Change Bundling Discrepancy  |  FLAG = service appears in oil-change visits but NOT in standalone non-oil visits  |  {date_window}',
        N_ID + n_c6)
    write_id_hdrs(ws6, 2)
    for col, cat in enumerate(DISC_CATS, N_ID+1):
        set_cat_hdr(ws6, 2, col, cat, rotation=60)
    ws6.row_dimensions[2].height = 80

    # Row 3: count of flagged stores per category
    set_count_label(ws6, 3, 1, N_ID, 'Stores with discrepancy (FLAG) →')
    flag_counts = defaultdict(int)
    for s in stores:
        for cat in DISC_CATS:
            has_all = s['l1_tx_all'].get(cat, 0) > 0
            has_non = s['l1_tx'].get(cat, 0) > 0
            if has_all and not has_non:
                flag_counts[cat] += 1
    for col, cat in enumerate(DISC_CATS, N_ID+1):
        c = ws6.cell(row=3, column=col, value=flag_counts[cat] or None)
        c.font = font(bold=True, color='92400E', size=10)
        c.fill = fill(FLAG_FILL if flag_counts[cat] else MGREY)
        c.alignment = align('center','center'); c.border = bdr()

    for i, s in enumerate(stores, 4):
        bg = LGREY if i % 2 == 0 else WHITE
        write_id_vals(ws6, i, s, bg)
        for col, cat in enumerate(DISC_CATS, N_ID+1):
            has_all = s['l1_tx_all'].get(cat, 0) > 0
            has_non = s['l1_tx'].get(cat, 0) > 0
            if has_all and not has_non:
                # FLAG: service only present when bundled with oil change
                c = ws6.cell(row=i, column=col, value='FLAG')
                c.fill = fill(FLAG_FILL)
                c.font = font(bold=True, color=FLAG_FONT, size=10)
            elif has_non:
                # Both: present in standalone non-oil visits too
                c = ws6.cell(row=i, column=col, value='Both')
                c.fill = fill(YES_FILL)
                c.font = font(bold=False, color=YES_FONT, size=10)
            else:
                # Neither
                c = ws6.cell(row=i, column=col, value='No')
                c.fill = fill(bg)
                c.font = font(color=DGREY, size=10)
            c.alignment = align('center','center'); c.border = bdr()
        ws6.row_dimensions[i].height = 15

    ws6.auto_filter.ref = f'A2:{get_column_letter(N_ID+n_c6)}{3+n_stores}'

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 7: Brand Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('Brand Summary')
    ws.freeze_panes = 'B4'
    n_brands = len(BRANDS)
    set_title(ws, 'FullSpeed Automotive — Service Penetration by Brand  |  % of brand stores with ≥1 transaction  |  Non-oil', 2+n_brands)
    c = ws.cell(row=2, column=1, value='Service Category')
    c.font = font(bold=True, color=WHITE, size=11); c.fill = fill(DARK2)
    c.alignment = align('left','center'); c.border = bdr()
    ws.column_dimensions['A'].width = 28
    for col, brand in enumerate(BRANDS, 2):
        c = ws.cell(row=2, column=col, value=brand)
        c.font = font(bold=True, color=WHITE, size=11); c.fill = fill(FSA_RED)
        c.alignment = align('center','bottom',rot=45); c.border = bdr()
        ws.column_dimensions[get_column_letter(col)].width = 14
    c = ws.cell(row=2, column=2+n_brands, value='All Brands')
    c.font = font(bold=True, color=WHITE, size=11); c.fill = fill(DARK)
    c.alignment = align('center','bottom',rot=45); c.border = bdr()
    ws.column_dimensions[get_column_letter(2+n_brands)].width = 14
    ws.row_dimensions[2].height = 80
    c = ws.cell(row=3, column=1, value='Store count →')
    c.font = font(bold=True, color=DGREY, size=9, italic=True)
    c.fill = fill(MGREY); c.alignment = align('right','center'); c.border = bdr()
    for col, brand in enumerate(BRANDS, 2):
        set_count_cell(ws, 3, col, len(brand_stores[brand]), FSA_RED)
    set_count_cell(ws, 3, 2+n_brands, n_stores, DARK)
    ws.row_dimensions[3].height = 16
    for i, l1 in enumerate(NON_OIL_CATS, 4):
        color = CAT_COLORS.get(l1,'6B7280')
        c = ws.cell(row=i, column=1, value=l1)
        c.font = font(bold=True, color=WHITE, size=11); c.fill = fill(color)
        c.alignment = align('left','center'); c.border = bdr()
        for j, brand in enumerate(BRANDS, 2):
            bs = len(brand_stores[brand]); os_ = len(brand_l1[brand][l1])
            pct = os_/bs if bs else 0
            c = ws.cell(row=i, column=j, value=hm_val(pct))
            c.fill = hm_fill(pct); c.font = hm_font(pct)
            c.alignment = align('center','center'); c.border = bdr()
        pct = len(l1_stats[l1]['stores'])/n_stores
        c = ws.cell(row=i, column=2+n_brands, value=hm_val(pct))
        c.fill = hm_fill(pct); c.font = hm_font(pct)
        c.alignment = align('center','center'); c.border = bdr()
        ws.row_dimensions[i].height = 18

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 8: Model Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('Model Summary')
    ws.freeze_panes = 'B4'
    set_title(ws, 'FullSpeed Automotive — Service Penetration by Store Model  |  Non-oil', 2+len(ALL_MODELS))
    c = ws.cell(row=2, column=1, value='Service Category')
    c.font = font(bold=True, color=WHITE, size=11); c.fill = fill(DARK2)
    c.alignment = align('left','center'); c.border = bdr(); ws.column_dimensions['A'].width = 28
    for col, m in enumerate(ALL_MODELS, 2):
        c = ws.cell(row=2, column=col, value=m)
        c.font = font(bold=True, color=WHITE, size=11); c.fill = fill(MODEL_COLORS[m])
        c.alignment = align('center','bottom',rot=45); c.border = bdr()
        ws.column_dimensions[get_column_letter(col)].width = 14
    c = ws.cell(row=2, column=2+len(ALL_MODELS), value='All Models')
    c.font = font(bold=True, color=WHITE, size=11); c.fill = fill(DARK)
    c.alignment = align('center','bottom',rot=45); c.border = bdr()
    ws.column_dimensions[get_column_letter(2+len(ALL_MODELS))].width = 14
    ws.row_dimensions[2].height = 80
    c = ws.cell(row=3, column=1, value='Store count →')
    c.font = font(bold=True, color=DGREY, size=9, italic=True)
    c.fill = fill(MGREY); c.alignment = align('right','center'); c.border = bdr()
    for col, m in enumerate(ALL_MODELS, 2):
        set_count_cell(ws, 3, col, len(model_stores.get(m,set())), MODEL_COLORS[m])
    set_count_cell(ws, 3, 2+len(ALL_MODELS), n_stores, DARK)
    ws.row_dimensions[3].height = 16
    for i, l1 in enumerate(NON_OIL_CATS, 4):
        color = CAT_COLORS.get(l1,'6B7280')
        c = ws.cell(row=i, column=1, value=l1)
        c.font = font(bold=True, color=WHITE, size=11); c.fill = fill(color)
        c.alignment = align('left','center'); c.border = bdr()
        for j, m in enumerate(ALL_MODELS, 2):
            ms = len(model_stores.get(m,set())); os_ = len(model_l1[m][l1])
            pct = os_/ms if ms else 0
            c = ws.cell(row=i, column=j, value=hm_val(pct))
            c.fill = hm_fill(pct); c.font = hm_font(pct)
            c.alignment = align('center','center'); c.border = bdr()
        pct = len(l1_stats[l1]['stores'])/n_stores
        c = ws.cell(row=i, column=2+len(ALL_MODELS), value=hm_val(pct))
        c.fill = hm_fill(pct); c.font = hm_font(pct)
        c.alignment = align('center','center'); c.border = bdr()
        ws.row_dimensions[i].height = 18

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 9: Region Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('Region Summary')
    ws.freeze_panes = 'B4'
    set_title(ws, 'FullSpeed Automotive — Service Penetration by Region  |  Non-oil', 2+len(REGIONS))
    c = ws.cell(row=2, column=1, value='Service Category')
    c.font = font(bold=True, color=WHITE, size=11); c.fill = fill(DARK2)
    c.alignment = align('left','center'); c.border = bdr(); ws.column_dimensions['A'].width = 28
    REG_COLORS = {'1':'0891B2','2':'059669','3':'7C3AED','4':'D97706'}
    for col, rn in enumerate(REGIONS, 2):
        vp = REGION_LABELS.get(rn,'')
        label = f"Region {rn}\n{vp}" if vp else f"Region {rn}"
        c = ws.cell(row=2, column=col, value=label)
        c.font = font(bold=True, color=WHITE, size=10)
        c.fill = fill(REG_COLORS.get(rn, DARK2))
        c.alignment = align('center','center',wrap=True); c.border = bdr()
        ws.column_dimensions[get_column_letter(col)].width = 18
    c = ws.cell(row=2, column=2+len(REGIONS), value='All Regions')
    c.font = font(bold=True, color=WHITE, size=11); c.fill = fill(DARK)
    c.alignment = align('center','center'); c.border = bdr()
    ws.column_dimensions[get_column_letter(2+len(REGIONS))].width = 14
    ws.row_dimensions[2].height = 36
    c = ws.cell(row=3, column=1, value='Store count →')
    c.font = font(bold=True, color=DGREY, size=9, italic=True)
    c.fill = fill(MGREY); c.alignment = align('right','center'); c.border = bdr()
    for col, rn in enumerate(REGIONS, 2):
        set_count_cell(ws, 3, col, len(region_stores[rn]), REG_COLORS.get(rn,DARK2))
    set_count_cell(ws, 3, 2+len(REGIONS), n_stores, DARK)
    ws.row_dimensions[3].height = 16
    for i, l1 in enumerate(NON_OIL_CATS, 4):
        color = CAT_COLORS.get(l1,'6B7280')
        c = ws.cell(row=i, column=1, value=l1)
        c.font = font(bold=True, color=WHITE, size=11); c.fill = fill(color)
        c.alignment = align('left','center'); c.border = bdr()
        for j, rn in enumerate(REGIONS, 2):
            rs = len(region_stores[rn]); os_ = len(region_l1[rn][l1])
            pct = os_/rs if rs else 0
            c = ws.cell(row=i, column=j, value=hm_val(pct))
            c.fill = hm_fill(pct); c.font = hm_font(pct)
            c.alignment = align('center','center'); c.border = bdr()
        pct = len(l1_stats[l1]['stores'])/n_stores
        c = ws.cell(row=i, column=2+len(REGIONS), value=hm_val(pct))
        c.fill = hm_fill(pct); c.font = hm_font(pct)
        c.alignment = align('center','center'); c.border = bdr()
        ws.row_dimensions[i].height = 18

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 10: DMA Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('DMA Summary')
    ws.freeze_panes = 'C4'
    set_title(ws, f'FullSpeed Automotive — Service Penetration by DMA  |  Non-oil  |  {date_window}', 2+len(NON_OIL_CATS))
    set_id_hdr(ws, 2, 1, 'DMA Market', 28); set_id_hdr(ws, 2, 2, 'Stores', 9)
    for col, cat in enumerate(NON_OIL_CATS, 3):
        set_cat_hdr(ws, 2, col, cat, rotation=60)
    ws.row_dimensions[2].height = 80
    ws.merge_cells('A3:B3')
    c = ws.cell(row=3, column=1, value='Stores with transactions →')
    c.font = font(bold=True, color=DGREY, size=9, italic=True)
    c.fill = fill(MGREY); c.alignment = align('right','center'); c.border = bdr()
    for col, cat in enumerate(NON_OIL_CATS, 3):
        set_count_cell(ws, 3, col, len(l1_stats[cat]['stores']), CAT_COLORS.get(cat,'6B7280'))
    ws.row_dimensions[3].height = 16
    for i, dma in enumerate(DMAS, 4):
        bg = LGREY if i % 2 == 0 else WHITE
        n = len(dma_stores[dma])
        c = ws.cell(row=i, column=1, value=dma)
        c.font = font(bold=True, size=10, color=BLACK); c.fill = fill(bg)
        c.alignment = align('left','center'); c.border = bdr()
        c = ws.cell(row=i, column=2, value=n)
        c.font = font(size=10, color=BLACK); c.fill = fill(bg)
        c.alignment = align('center','center'); c.border = bdr()
        for col, cat in enumerate(NON_OIL_CATS, 3):
            os_ = len(dma_l1[dma][cat]); pct = os_/n if n else 0
            c = ws.cell(row=i, column=col, value=hm_val(pct))
            c.fill = hm_fill(pct); c.font = hm_font(pct)
            c.alignment = align('center','center'); c.border = bdr()
        ws.row_dimensions[i].height = 15
    ws.auto_filter.ref = f'A2:{get_column_letter(2+len(NON_OIL_CATS))}{3+len(DMAS)}'

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 11: Hierarchy Reference
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('Hierarchy Reference')
    ws.freeze_panes = 'A3'
    set_title(ws, 'Product Hierarchy — L1 Categories & L2 Subcategories  |  Coverage statistics  |  v2', 6)
    for col, (hdr, w) in enumerate([('L1 Category',28),('L2 Subcategory',30),
                                     ('Total Transactions',16),('Stores Offering',14),
                                     ('% of All Stores',14),('Change Notes',42)], 1):
        set_id_hdr(ws, 2, col, hdr, w)
    ws.row_dimensions[2].height = 18
    rn = 3
    for l1 in NON_OIL_CATS:
        color = CAT_COLORS.get(l1,'6B7280')
        l2s = AUTHORITATIVE_L2[l1]
        l1_s = len(l1_stats[l1]['stores']); l1_tx = l1_stats[l1]['tx']
        for j, l2 in enumerate(l2s):
            bg = LGREY if rn % 2 == 0 else WHITE
            stats = l2_stats.get((l1,l2), {'stores':set(),'tx':0})
            c = ws.cell(row=rn, column=1, value=l1 if j==0 else '')
            c.font = font(bold=True,color=WHITE,size=11) if j==0 else font(size=10,color=BLACK)
            c.fill = fill(color) if j==0 else fill(bg)
            c.alignment = align('left','center'); c.border = bdr()
            for col, val in [(2,l2),(3,stats['tx']),(4,len(stats['stores'])),(5,len(stats['stores'])/n_stores if n_stores else 0),(6,'')]:
                c = ws.cell(row=rn, column=col, value=val)
                c.font = font(size=10); c.fill = fill(bg)
                c.alignment = align('center' if col>2 else 'left', 'center'); c.border = bdr()
                if col==3: c.number_format='#,##0'
                if col==5: c.number_format='0%'
            rn += 1
        c = ws.cell(row=rn, column=1, value=f'TOTAL — {l1}')
        c.font = font(bold=True,color=WHITE,size=11); c.fill = fill(DARK)
        c.alignment = align('left','center'); c.border = bdr()
        for col, val, fmt in [(2,'',None),(3,l1_tx,'#,##0'),(4,l1_s,None),(5,l1_s/n_stores if n_stores else 0,'0%'),(6,'',None)]:
            c = ws.cell(row=rn, column=col, value=val)
            c.font = font(bold=True,color=WHITE,size=11); c.fill = fill(DARK)
            c.alignment = align('center','center'); c.border = bdr()
            if fmt: c.number_format = fmt
        rn += 1

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 12: Raw Data
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('Raw Data')
    ws.freeze_panes = 'A3'
    set_title(ws, f'Raw Data — All Store × Service Combinations  |  {len(rows):,} rows  |  {date_window}', 14)
    RCOLS = [('STORE_NUMBER',9),('STORE_NAME',30),('BRAND',22),('CITY',14),('STATE',6),
             ('DMA_NAME',26),('REGION_NUM',8),('DISTRICT_NUM',9),
             ('STORE_MODEL',10),('PROPOSED_MODEL',13),
             ('L1_CATEGORY',22),('L2_SUBCATEGORY',24),
             ('SERVICE_NAME',34),('TRANSACTION_COUNT',14),('IS_OIL_TRANSACTION',14),('IS_INSPECTION_STATE',16)]
    for col, (hdr, w) in enumerate(RCOLS, 1):
        set_id_hdr(ws, 2, col, hdr, w)
    ws.row_dimensions[2].height = 18
    for i, r in enumerate(rows, 3):
        bg = LGREY if i % 2 == 0 else WHITE
        for col, (key, _) in enumerate(RCOLS, 1):
            val = r.get(key, '')
            if key == 'TRANSACTION_COUNT' and val:
                try: val = int(val)
                except: pass
            if key == 'IS_OIL_TRANSACTION':
                val = 'Yes' if str(val).strip() == '1' else 'No'
            c = ws.cell(row=i, column=col, value=val)
            c.font = font(size=9, color=BLACK); c.fill = fill(bg)
            c.alignment = align('left','center'); c.border = bdr()
        ws.row_dimensions[i].height = 14
    ws.auto_filter.ref = f'A2:{get_column_letter(len(RCOLS))}{2+len(rows)}'

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 13: Store Directory
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('Store Directory')
    ws.freeze_panes = 'A4'
    set_title(ws, f'Store Directory  |  {n_stores} Active Locations  |  {date_window}', 12)
    ws.merge_cells('A2:L2')
    c = ws.cell(row=2, column=1, value='All active locations with brand, geography, DMA, region, district, and store model tier')
    c.font = font(color=DGREY, size=10, italic=True); c.fill = fill(MGREY); c.alignment = align('left','center')
    ws.row_dimensions[2].height = 16
    for col, (hdr, w) in enumerate([('Store #',9),('Brand',22),('Store Name',30),
                                     ('City',16),('State',7),('Zip',8),('DMA',26),
                                     ('Region',8),('District',9),('District Mgr',20),
                                     ('Region VP',20),('Model',9)], 1):
        set_id_hdr(ws, 3, col, hdr, w)
    ws.row_dimensions[3].height = 18
    for i, s in enumerate(stores, 4):
        bg = LGREY if i % 2 == 0 else WHITE
        for col, val in enumerate([s['STORE_NUMBER'],s['BRAND'],s['STORE_NAME'],
                                    s['CITY'],s['STATE'],s['ZIP_CODE'],s['DMA_NAME'],
                                    s['REGION_NUM'],s['DISTRICT_NUM'],s['DISTRICT_MGR'],
                                    s['REGION_VP'],s['STORE_MODEL']], 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = font(size=10, color=BLACK); c.fill = fill(bg)
            c.alignment = align('left','center'); c.border = bdr()
        ws.row_dimensions[i].height = 15
    ws.auto_filter.ref = f'A3:L{3+n_stores}'

    # ── SAVE ───────────────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"  Written: {output_path}")
    print(f"  Sheets: {[s.title for s in wb.worksheets]}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--input',           required=True)
    parser.add_argument('--output',          required=True)
    parser.add_argument('--region-district', default='scripts/region_district.csv')
    args = parser.parse_args()
    build(args.input, args.output, args.region_district)

if __name__ == '__main__':
    main()
